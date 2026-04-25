"""Auto-fit openpyxl column widths to content.

Why: openpyxl has no built-in column-autosize, and the recipes you find online
either ignore formatted numbers, blow up on merged cells, or set widths to
the longest value with no upper bound (so a 500-char description column makes
the sheet unreadable). This is the practical default I keep rewriting.

Usage:
    from openpyxl import load_workbook
    wb = load_workbook("input.xlsx")
    autosize_columns(wb["Sheet1"])
    wb.save("output.xlsx")

Deps: pip install openpyxl
"""
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

MIN_WIDTH = 8
MAX_WIDTH = 60
PADDING = 2


def autosize_columns(
    ws: Worksheet,
    *,
    min_width: int = MIN_WIDTH,
    max_width: int = MAX_WIDTH,
    padding: int = PADDING,
) -> None:
    """Set column widths to fit the longest cell value in each column."""
    widths: dict[str, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.column_letter is None or cell.value is None:
                continue
            text = str(cell.value)
            current = widths.get(cell.column_letter, 0)
            widths[cell.column_letter] = max(current, len(text))
    for col, width in widths.items():
        bounded = max(min_width, min(width + padding, max_width))
        ws.column_dimensions[col].width = bounded
